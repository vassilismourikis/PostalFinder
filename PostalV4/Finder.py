import openpyxl
import xlrd
from itertools import takewhile
import difflib
import time

def trans(str):
    greek_alphabet = 'ΑαΒβΓγΔδΕεΖζΗηΘθΙιΚκΛλΜμΝνΞξΟοΠπΡρΣσΤτΥυΦφΧχΨψΩω'
    latin_alphabet = 'AaBbGgDdEeZzHhJjIiKkLlMmNnXxOoPpRrSsTtUuFfQqYyWw'
    latin2greek = str.maketrans(latin_alphabet,greek_alphabet)

    return str.translate(latin2greek)

# ----------------------------------------------------------------------
#ONLY xlsx files because xls are old and are not supported from openpyxl
def write_results(path,data):
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.open(path)

    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active

    # Note: The first row or column integer
    # is 1, not 0. Cell object is created by
    # using sheet object's cell() method.
    for index,i in enumerate(data):
        c1 = sheet.cell(index+2, 10)

        # writing values to cells
        c1.value = i

    # Anytime you modify the Workbook object
    # or its sheets and cells, the spreadsheet
    # file will not be saved until you call
    # the save() workbook method.
    wb.save(path)
    wb.close()
#---------------------------------------------------------------------------

#GET COLUMN NUMBER
def column_len(sheet, index):
    col_values = sheet.col_values(index)
    col_len = len(col_values)
    for _ in takewhile(lambda x: not x, reversed(col_values)):
        col_len -= 1
    return col_len

#GET COLUMN'S DATA
def get_data(path,until):
    names=[]
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    for i in range(1,column_len(sheet, 1)):
        names.append(str(sheet.cell(i,until).value))
    return names


# ----------------------------------------------------------------------
if __name__ == "__main__":
    print("Loading Database...")
    path1 ="attiki.xls"
    path2="thess.xls"
    path3="loipi.xls"
    tks = get_data(path1, 2)
    tks.extend(get_data(path2, 2))
    tks.extend(get_data(path3, 2))
    streets=get_data(path1,0)
    streets.extend(get_data(path2, 0))
    streets.extend(get_data(path3, 0))
    towns=get_data(path1,3)
    towns.extend(get_data(path2, 3))
    towns.extend(get_data(path3, 3))
    empty_path = input("Give file's path: ")
    print("Loading file's data...")
    forSearchAdress=get_data(empty_path,4)
    forSearchTown=get_data(empty_path,2)
    alterForSearchTown = get_data(empty_path, 3)


    #EXTRACTING ACTUAL TOWN FROM DB DATA
    for index,y in enumerate(forSearchTown):
        t=''
        o=False
        #when they chosing wrong town but they typing it
        if(y=="Η πόλη"):
            forSearchTown[index] = trans(alterForSearchTown[index])
            continue
        for c in y:
            if((c=='-')&(t=='')):
                o=True
                continue
            elif((c=='-')&(t!='')):
                o=False
            if(o):
                t+=c
        forSearchTown[index]=t

    print("Searching...")

    foundAdresses=[]

    #TOWN VALIDATION
    for outIndex,i in enumerate(forSearchAdress):
        k=difflib.get_close_matches(trans(i).upper(), streets, 1)
        #if address is available
        if(k!=[]):

            t = difflib.get_close_matches(trans(forSearchTown[outIndex]).upper(), towns, 1)
            for index,j in enumerate(streets):
                fLen=len(foundAdresses)
                #if town is available
                if(t!=[]):
                    if ((difflib.SequenceMatcher(None, k[0], j).ratio() > 0.75) & (
                            difflib.SequenceMatcher(None, t[0], towns[index]).ratio() > 0.75)):
                            foundAdresses.append(tks[index])
                            break
                else:
                    break
            #if combination of town and address didnt found search with the alternative town
            if(len(foundAdresses)==fLen):
                t = difflib.get_close_matches(trans(alterForSearchTown[outIndex]).upper(), towns, 1)
                for index, j in enumerate(streets):
                    fLen = len(foundAdresses)
                    if (t != []):

                        if((difflib.SequenceMatcher(None, k[0], j).ratio() > 0.75) & (difflib.SequenceMatcher(None, t[0], towns[index]).ratio()>0.75)):
                                foundAdresses.append(tks[index])
                                break
                    else:
                        break
                #PRINTING FINAL NOT FOUNDS
                if (len(foundAdresses) == fLen):
                    mult_tks=''
                    #APPENDING THE CELL WITH ALL THE TKS FROM DIFFERENT TOWNS TO DO IT BY HAND
                    #IF ADDRESS FOUND BEFORE
                    for index,j in enumerate(streets):
                        if(difflib.SequenceMatcher(None, k[0], j).ratio() > 0.90):
                            mult_tks+=str(tks[index]+',')
                    foundAdresses.append(mult_tks)

        #if address didnt found
        else:
            #PRINTING THESE THAT DIDN'T EVEN FOUND THEIR ADRESS
            #print(i)
            foundAdresses.append(None)

    #COUNTING FOUNDS
    i=0
    for t in foundAdresses:
        if t!=None:
            i+=1
    print("Found " + str(i) + " results")

    print("writing to excel,finishing up...")
    write_results(empty_path,foundAdresses)
    print("All ready stopping execution...")
    time.sleep(5)

